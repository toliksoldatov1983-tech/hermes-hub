from __future__ import annotations

import importlib.util
import sys
import types
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
MODULE_PATH = ROOT / "malyarka_file.py"

RUSSIAN_HEADERS = [
    "№",
    "Высота",
    "Ширина",
    "Количество",
    "Площадь детали, м²",
    "Площадь строки, м²",
    "Статус",
    "Исходный текст",
    "Причина спора",
]

FORBIDDEN_ENGLISH = {
    "row_number",
    "height",
    "width",
    "quantity",
    "item_area",
    "row_area",
    "row_status",
    "raw_text",
    "dispute_reason",
    "TOTAL_CONFIRMED",
    "confirmed",
    "confirmed_total",
}


@dataclass
class FakeOrderItem:
    height: int
    width: int
    quantity: int
    source: str = ""


@dataclass
class FakeDisputedItem:
    raw: str
    reason: str


@dataclass
class FakeOrderDraft:
    items: list[FakeOrderItem] = field(default_factory=list)
    disputed_items: list[FakeDisputedItem] = field(default_factory=list)


def install_fake_malyarka_core() -> None:
    core = types.ModuleType("malyarka_core")
    models = types.ModuleType("malyarka_core.models")
    validation = types.ModuleType("malyarka_core.validation")

    models.OrderItem = FakeOrderItem
    models.DisputedItem = FakeDisputedItem
    models.OrderDraft = FakeOrderDraft

    def validate_order_for_export(order):
        if not order.items or order.disputed_items:
            raise ValueError("not exportable")

    validation.validate_order_for_export = validate_order_for_export

    sys.modules["malyarka_core"] = core
    sys.modules["malyarka_core.models"] = models
    sys.modules["malyarka_core.validation"] = validation


def load_candidate_module():
    install_fake_malyarka_core()
    spec = importlib.util.spec_from_file_location("candidate_malyarka_file", MODULE_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def test_malyarka_file_xlsx_uses_russian_headers_and_statuses(tmp_path):
    module = load_candidate_module()
    order = FakeOrderDraft(items=[FakeOrderItem(700, 500, 1, source="700 x 500")])
    output_path = tmp_path / "malyarka_file.xlsx"

    module.export_malyarka_file_xlsx(order, output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook.active
    values = [
        cell
        for row in worksheet.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    ]

    assert [worksheet.cell(row=1, column=index).value for index in range(1, 10)] == RUSSIAN_HEADERS
    assert "подтверждено" in values
    assert "итого подтверждено" in values
    assert "ИТОГО ПОДТВЕРЖДЕНО" in values
    assert not (FORBIDDEN_ENGLISH & {str(value) for value in values})
