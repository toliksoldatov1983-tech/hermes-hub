"""Tests for BATCH_114: Staging File Creation."""

from __future__ import annotations
import sys
from pathlib import Path
import pytest

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT / "src") not in sys.path:
    sys.path.insert(0, str(ROOT / "src"))

STAGING = ROOT / "06_EXPORT_STAGING"


class TestStagingFilesExist:
    def test_staging_folder_exists(self):
        assert STAGING.exists(), "Staging folder should exist"

    def test_corel_txt_exists(self):
        assert (STAGING / "demo_order_corel.txt").exists()

    def test_excel_xlsx_exists(self):
        assert (STAGING / "demo_order_malyarka.xlsx").exists()

    def test_preview_json_exists(self):
        assert (STAGING / "demo_order_export_preview.json").exists()

    def test_report_md_exists(self):
        assert (STAGING / "demo_order_export_report.md").exists()


class TestCorelTXTContent:
    def test_first_line_empty(self):
        content = (STAGING / "demo_order_corel.txt").read_text()
        assert content.startswith("\n") or content.startswith("\r\n")

    def test_tab_delimiter(self):
        content = (STAGING / "demo_order_corel.txt").read_text()
        assert "\t" in content

    def test_no_headers(self):
        content = (STAGING / "demo_order_corel.txt").read_text().lower()
        assert "height" not in content
        assert "width" not in content

    def test_numeric_values(self):
        content = (STAGING / "demo_order_corel.txt").read_text()
        for line in content.strip().splitlines():
            if line.strip():
                parts = line.split("\t")
                assert len(parts) == 3
                for p in parts:
                    assert p.strip().isdigit()


class TestExcelContent:
    def test_headers(self):
        import openpyxl
        wb = openpyxl.load_workbook(str(STAGING / "demo_order_malyarka.xlsx"))
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, 10)]
        assert "H" in headers
        assert "м²" in headers

    def test_area_cell(self):
        import openpyxl
        wb = openpyxl.load_workbook(str(STAGING / "demo_order_malyarka.xlsx"))
        ws = wb.active
        assert isinstance(ws.cell(2, 5).value, (int, float))


class TestSafety:
    def test_all_files_in_staging(self):
        assert all(f.resolve().is_relative_to(STAGING.resolve()) for f in STAGING.glob("*") if f.is_file())

    def test_no_e_disk_write(self):
        # E:\Заказы should not have demo files
        e_zakazy = Path("E:/Заказы")
        if e_zakazy.exists():
            demo_files = list(e_zakazy.glob("demo_order_*"))
            assert len(demo_files) == 0, f"E:\\Заказы should not have demo files: {demo_files}"


class TestRegression:
    def test_malyarka_fixtures(self):
        from hermes_modules.malyarka.fixtures import run_all_fixtures
        assert len(run_all_fixtures()) == 12

    def test_test_count(self):
        import subprocess
        result = subprocess.run([sys.executable, "-m", "pytest", "tests/", "--collect-only", "-q"],
                                capture_output=True, text=True, cwd=str(ROOT))
        assert "tests collected" in result.stdout
