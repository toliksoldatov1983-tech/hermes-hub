import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from malyarka_core.exports.corel import build_corel_rows
from malyarka_core.exports.malyarka_file import (
    MALYARKA_FILE_HEADERS,
    build_malyarka_file_rows,
    build_malyarka_file_summary,
    build_malyarka_file_table,
    calculate_row_area_m2,
    calculate_single_item_area_m2,
    export_malyarka_file_xlsx,
)
from malyarka_core.models import DisputedItem, OrderDraft, OrderItem
from malyarka_core.parsing import parse_sizes_text


def test_malyarka_file_rows_include_confirmed_area_and_raw_text():
    order = parse_sizes_text("1000*400*2\n500 700")

    rows = build_malyarka_file_rows(order)

    assert len(rows) == 2
    assert rows[0].row_number == 1
    assert rows[0].height == 1000
    assert rows[0].width == 400
    assert rows[0].quantity == 2
    assert rows[0].item_area_m2 == pytest.approx(0.4)
    assert rows[0].row_area_m2 == pytest.approx(0.8)
    assert rows[0].row_status == "confirmed"
    assert rows[0].raw_text == "1000*400*2"

    assert rows[1].item_area_m2 == pytest.approx(0.35)
    assert rows[1].row_area_m2 == pytest.approx(0.35)


def test_malyarka_file_summary_counts_only_confirmed_area():
    order = OrderDraft(
        items=[
            OrderItem(height=1000, width=400, quantity=2, source="1000 400 2"),
            OrderItem(height=500, width=700, quantity=1, source="500 700"),
        ],
        disputed_items=[
            DisputedItem(raw="непонятная строка", reason="Найдена спорная строка.")
        ],
    )

    summary = build_malyarka_file_summary(order)

    assert summary.total_confirmed_quantity == 3
    assert summary.total_confirmed_area_m2 == pytest.approx(1.15)
    assert summary.confirmed_count == 2
    assert summary.disputed_count == 1
    assert summary.can_export_final_file is False


def test_malyarka_file_rows_include_disputed_status_and_reason():
    order = parse_sizes_text("1000*400\nнепонятная строка")

    rows = build_malyarka_file_rows(order)

    assert rows[1].row_number == 2
    assert rows[1].height is None
    assert rows[1].width is None
    assert rows[1].quantity is None
    assert rows[1].item_area_m2 is None
    assert rows[1].row_area_m2 is None
    assert rows[1].row_status == "disputed"
    assert rows[1].raw_text == "непонятная строка"
    assert rows[1].dispute_reason


def test_malyarka_file_table_has_minimal_headers_and_confirmed_total():
    order = parse_sizes_text("1000*400*2\n500 700")

    table = build_malyarka_file_table(order)

    assert table[0] == MALYARKA_FILE_HEADERS
    assert table[1] == [
        1,
        1000,
        400,
        2,
        pytest.approx(0.4),
        pytest.approx(0.8),
        "подтверждено",
        "1000*400*2",
        None,
    ]
    assert table[-1] == [
        "ИТОГО ПОДТВЕРЖДЕНО",
        None,
        None,
        3,
        None,
        pytest.approx(1.15),
        "итого подтверждено",
        "",
        None,
    ]


def test_malyarka_file_export_xlsx_for_clean_order(tmp_path):
    order = parse_sizes_text("1000*400*2\n500 700")
    output_path = tmp_path / "malyarka_file.xlsx"

    result = export_malyarka_file_xlsx(order, output_path)

    assert result == output_path
    workbook = load_workbook(output_path, data_only=True)
    worksheet = workbook.active
    assert worksheet.title == "Malyarka File"
    assert [worksheet.cell(1, column).value for column in range(1, 10)] == (
        MALYARKA_FILE_HEADERS
    )
    assert worksheet["A2"].value == 1
    assert worksheet["B2"].value == 1000
    assert worksheet["C2"].value == 400
    assert worksheet["D2"].value == 2
    assert worksheet["E2"].value == pytest.approx(0.4)
    assert worksheet["F2"].value == pytest.approx(0.8)
    assert worksheet["G2"].value == "подтверждено"
    assert worksheet["H2"].value == "1000*400*2"
    assert worksheet["A4"].value == "ИТОГО ПОДТВЕРЖДЕНО"
    assert worksheet["D4"].value == 3
    assert worksheet["F4"].value == pytest.approx(1.15)


def test_malyarka_file_export_blocks_disputed_order(tmp_path):
    order = parse_sizes_text("1000*400\nнепонятная строка")
    output_path = tmp_path / "malyarka_file.xlsx"

    with pytest.raises(ValueError):
        export_malyarka_file_xlsx(order, output_path)

    assert not output_path.exists()
    summary = build_malyarka_file_summary(order)
    assert summary.total_confirmed_area_m2 == pytest.approx(0.4)
    assert summary.can_export_final_file is False


def test_malyarka_file_empty_order_is_not_exportable(tmp_path):
    order = OrderDraft()
    output_path = tmp_path / "empty_malyarka_file.xlsx"

    assert build_malyarka_file_rows(order) == []
    summary = build_malyarka_file_summary(order)
    assert summary.total_confirmed_quantity == 0
    assert summary.total_confirmed_area_m2 == 0
    assert summary.can_export_final_file is False
    with pytest.raises(ValueError):
        export_malyarka_file_xlsx(order, output_path)
    assert not output_path.exists()


def test_malyarka_file_area_helpers_use_required_formulas():
    item = OrderItem(height=1200, width=600, quantity=3)

    assert calculate_single_item_area_m2(item) == pytest.approx(0.72)
    assert calculate_row_area_m2(item) == pytest.approx(2.16)


def test_corel_export_rows_stay_height_width_quantity_only():
    order = parse_sizes_text("1000*400*2\n500 700")

    assert build_corel_rows(order) == [[], [1000, 400, 2], [500, 700, 1]]
