import asyncio
from pathlib import Path
from types import SimpleNamespace
from zipfile import ZipFile

from openpyxl import load_workbook

from malyarka_telegram.app import (
    build_live_text_response,
    get_runtime_session_store,
    handle_clarify_intent_callback,
    handle_corel_excel_callback,
    handle_engineer_task_callback,
    handle_malyarka_file_callback,
    prepare_corel_excel_for_user,
    prepare_malyarka_file_for_user,
    resolve_message_user_id,
    split_engineer_approval_message,
    split_telegram_text,
)
from malyarka_telegram.config import TelegramConfig
from malyarka_telegram.handlers import handle_text_message
from malyarka_telegram.modes import TelegramMode
from malyarka_telegram.session import InMemoryModeSessionStore


def test_app_import_does_not_start_polling():
    import malyarka_telegram.app as app_module

    assert hasattr(app_module, "build_live_text_response")
    assert "Bot" not in vars(app_module)
    assert "Dispatcher" not in vars(app_module)


def test_app_has_runtime_in_memory_session_store():
    store = get_runtime_session_store()

    assert isinstance(store, InMemoryModeSessionStore)


def test_app_source_uses_routed_handler_for_live_text():
    source = Path("malyarka_telegram/app.py").read_text(encoding="utf-8")

    assert "handle_text_message_with_router" in source
    assert "build_text_response(text" not in source


def test_resolve_message_user_id_uses_from_user_id():
    message = SimpleNamespace(from_user=SimpleNamespace(id=777))

    assert resolve_message_user_id(message) == 777


def test_resolve_message_user_id_falls_back_safely():
    assert resolve_message_user_id(SimpleNamespace(from_user=None)) == 0
    assert resolve_message_user_id(SimpleNamespace()) == 0


def test_order_command_switches_mode_without_parsing():
    store = InMemoryModeSessionStore()

    response = build_live_text_response(
        text="/заказ",
        user_id=1,
        session_store=store,
    )

    assert store.get_mode(1) == TelegramMode.ORDER
    assert "Включён режим заказа" in response.text
    assert response.copy_keyboard is None


def test_order_mode_sends_text_to_preview_and_keeps_excel_and_copy_buttons():
    store = InMemoryModeSessionStore()
    build_live_text_response(text="/заказ", user_id=1, session_store=store)

    response = build_live_text_response(
        text="1000*400",
        user_id=1,
        session_store=store,
    )

    assert "Предпросмотр заказа" in response.text
    assert "1. 1000 400 1" in response.text
    assert response.copy_keyboard is not None
    assert [button.label for button in response.copy_keyboard.buttons] == [
        "Скачать Excel для Corel",
        "Скачать Файл Малярки",
        " Скопировать для Corel",
    ]
    assert response.copy_keyboard.buttons[0].action == "download_corel_excel"
    assert response.copy_keyboard.buttons[1].action == "download_malyarka_file"
    assert response.copy_keyboard.buttons[2].copy_text == "1000 400 1"


def test_live_order_context_prepares_corel_excel_in_temp_dir(tmp_path):
    store = InMemoryModeSessionStore()
    build_live_text_response(text="/заказ", user_id=10, session_store=store)
    build_live_text_response(
        text="1000*400\n1000*600*2",
        user_id=10,
        session_store=store,
    )

    result = prepare_corel_excel_for_user(user_id=10, output_dir=tmp_path)

    assert result.action == "download_corel_excel"
    assert result.sent_to_telegram is False
    assert result.path.parent == tmp_path
    assert result.path.suffix == ".xlsx"
    assert result.path.exists()
    assert result.rows_count == 2
    with ZipFile(result.path) as archive:
        sheet = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert '<c r="A2"><v>1000</v></c>' in sheet
    assert '<c r="B2"><v>400</v></c>' in sheet
    assert '<c r="C2"><v>1</v></c>' in sheet
    assert '<c r="A3"><v>1000</v></c>' in sheet
    assert '<c r="B3"><v>600</v></c>' in sheet
    assert '<c r="C3"><v>2</v></c>' in sheet


def test_live_order_context_prepares_malyarka_file_in_temp_dir(tmp_path):
    store = InMemoryModeSessionStore()
    build_live_text_response(text="/заказ", user_id=15, session_store=store)
    build_live_text_response(
        text="1000*400\n1000*600*2",
        user_id=15,
        session_store=store,
    )

    result = prepare_malyarka_file_for_user(user_id=15, output_dir=tmp_path)

    assert result.action == "download_malyarka_file"
    assert result.sent_to_telegram is False
    assert result.path.parent == tmp_path
    assert result.path.suffix == ".xlsx"
    assert result.path.exists()
    assert result.rows_count == 2
    workbook = load_workbook(result.path, data_only=True)
    worksheet = workbook.active
    assert worksheet["A1"].value == "№"
    assert worksheet["G2"].value == "подтверждено"
    assert worksheet["F2"].value == 0.4
    assert worksheet["A4"].value == "ИТОГО ПОДТВЕРЖДЕНО"
    assert worksheet["F4"].value == 1.6


def test_live_order_context_prepares_corel_excel_when_copy_button_is_skipped(tmp_path):
    long_order = "\n".join(f"{100 + index} 500 1" for index in range(30))
    store = InMemoryModeSessionStore()
    build_live_text_response(text="/заказ", user_id=12, session_store=store)
    build_live_text_response(text=long_order, user_id=12, session_store=store)

    result = prepare_corel_excel_for_user(user_id=12, output_dir=tmp_path)

    assert result.sent_to_telegram is False
    assert result.rows_count == 30
    with ZipFile(result.path) as archive:
        sheet = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert '<c r="A2"><v>100</v></c>' in sheet
    assert '<c r="A31"><v>129</v></c>' in sheet
    assert '<c r="A32">' not in sheet


def test_corel_excel_callback_sends_prepared_document_for_clean_order(tmp_path):
    store = InMemoryModeSessionStore()
    build_live_text_response(text="/заказ", user_id=13, session_store=store)
    build_live_text_response(text="1000*400\n1000*600*2", user_id=13, session_store=store)
    callback = FakeCallback(user_id=13)

    result = asyncio.run(
        handle_corel_excel_callback(
            callback,
            output_dir=tmp_path,
            document_factory=lambda path: path,
        )
    )

    assert result.sent_to_telegram is True
    assert result.path.parent == tmp_path
    assert result.path.suffix == ".xlsx"
    assert callback.answers == [
        {"text": "Excel для Corel готов. Отправляю файл.", "show_alert": False}
    ]
    assert len(callback.message.documents) == 1
    document_call = callback.message.documents[0]
    assert document_call["caption"] == "Excel для Corel"
    assert Path(document_call["document"]).parent == tmp_path
    assert Path(document_call["document"]).name == result.path.name
    with ZipFile(document_call["document"]) as archive:
        sheet = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert '<c r="A2"><v>1000</v></c>' in sheet
    assert '<c r="B3"><v>600</v></c>' in sheet
    assert '<c r="C3"><v>2</v></c>' in sheet


def test_malyarka_file_callback_sends_prepared_document_for_clean_order(tmp_path):
    store = InMemoryModeSessionStore()
    build_live_text_response(text="/заказ", user_id=16, session_store=store)
    build_live_text_response(text="1000*400\n1000*600*2", user_id=16, session_store=store)
    callback = FakeCallback(user_id=16, data="download_malyarka_file")

    result = asyncio.run(
        handle_malyarka_file_callback(
            callback,
            output_dir=tmp_path,
            document_factory=lambda path: path,
        )
    )

    assert result.sent_to_telegram is True
    assert result.path.parent == tmp_path
    assert result.path.suffix == ".xlsx"
    assert callback.answers == [
        {"text": "Файл Малярки готов. Отправляю файл.", "show_alert": False}
    ]
    assert len(callback.message.documents) == 1
    document_call = callback.message.documents[0]
    assert document_call["caption"] == "Файл Малярки"
    workbook = load_workbook(document_call["document"], data_only=True)
    worksheet = workbook.active
    assert worksheet["A1"].value == "№"
    assert worksheet["G2"].value == "подтверждено"
    assert worksheet["A4"].value == "ИТОГО ПОДТВЕРЖДЕНО"


def test_corel_excel_callback_does_not_send_document_for_disputed_order(tmp_path):
    store = InMemoryModeSessionStore()
    build_live_text_response(text="/заказ", user_id=14, session_store=store)
    build_live_text_response(text="1000*400\nнепонятная строка", user_id=14, session_store=store)
    callback = FakeCallback(user_id=14)

    result = asyncio.run(handle_corel_excel_callback(callback, output_dir=tmp_path))

    assert result is None
    assert callback.message.documents == []
    assert callback.answers == [
        {
            "text": "Нет подготовленного чистого заказа для Excel Corel.",
            "show_alert": True,
        }
    ]


def test_malyarka_file_callback_does_not_send_document_for_disputed_order(tmp_path):
    store = InMemoryModeSessionStore()
    build_live_text_response(text="/заказ", user_id=17, session_store=store)
    build_live_text_response(text="1000*400\nнепонятная строка", user_id=17, session_store=store)
    callback = FakeCallback(user_id=17, data="download_malyarka_file")

    result = asyncio.run(handle_malyarka_file_callback(callback, output_dir=tmp_path))

    assert result is None
    assert callback.message.documents == []
    assert callback.answers == [
        {
            "text": "Нет подготовленного чистого заказа для Файла Малярки.",
            "show_alert": True,
        }
    ]


def test_engineer_task_allow_callback_is_read_only():
    store = InMemoryModeSessionStore()
    build_live_text_response(
        text="/инженер",
        user_id=21,
        session_store=store,
        config=TelegramConfig(owner_id=21),
    )
    build_live_text_response(
        text="Сделать первый кодовый этап Файла Малярки.",
        user_id=21,
        session_store=store,
        config=TelegramConfig(owner_id=21),
    )
    callback = FakeCallback(user_id=21, data="engineer_task_allow_read_only")

    result = asyncio.run(
        handle_engineer_task_callback(callback, config=TelegramConfig(owner_id=21))
    )

    assert result is not None
    assert "ЗАДАНИЕ ДЛЯ CODEX/HERMES" in result
    assert "ПАКЕТ ДЛЯ НЕЗАВИСИМОГО РЕВЬЮ" in result
    assert "Сделать первый кодовый этап Файла Малярки." in result
    assert "Не запускать Hermes/Codex из Telegram" in result
    assert "Не вызывать DeepSeek/OpenAI/API" in result
    assert callback.answers == [
        {
            "text": "Задание и пакет ревью подготовлены для ручного копирования.",
            "show_alert": False,
        }
    ]
    assert "\n\n".join(callback.message.texts) == result
    assert all(len(text) <= 3900 for text in callback.message.texts)
    assert len(callback.message.texts) == 2
    assert callback.message.texts[0].startswith("ЗАДАНИЕ ДЛЯ CODEX/HERMES")
    assert callback.message.texts[1].startswith("ПАКЕТ ДЛЯ НЕЗАВИСИМОГО РЕВЬЮ")
    assert callback.message.documents == []


def test_split_telegram_text_keeps_short_text_and_splits_long_text():
    short = "short text"
    long_text = "A" * 2000 + "\n\n" + "B" * 2000 + "\n\n" + "C" * 2000

    assert split_telegram_text(short, limit=20) == [short]
    chunks = split_telegram_text(long_text, limit=2500)

    assert len(chunks) == 3
    assert all(len(chunk) <= 2500 for chunk in chunks)
    assert "\n\n".join(chunks) == long_text


def test_split_engineer_approval_message_separates_assignment_and_review():
    text = "ЗАДАНИЕ ДЛЯ CODEX/HERMES\n\nbody\n\nПАКЕТ ДЛЯ НЕЗАВИСИМОГО РЕВЬЮ\n\nreview"

    chunks = split_engineer_approval_message(text)

    assert chunks == [
        "ЗАДАНИЕ ДЛЯ CODEX/HERMES\n\nbody",
        "ПАКЕТ ДЛЯ НЕЗАВИСИМОГО РЕВЬЮ\n\nreview",
    ]


def test_engineer_task_cancel_callback_is_read_only():
    callback = FakeCallback(user_id=22, data="engineer_task_cancel_read_only")

    result = asyncio.run(
        handle_engineer_task_callback(callback, config=TelegramConfig(owner_id=22))
    )

    assert result is not None
    assert "отмен" in result.lower()
    assert "Никакие действия не выполнялись" in result
    assert callback.answers == [{"text": result, "show_alert": False}]
    assert callback.message.documents == []


def test_engineer_task_edit_callback_is_read_only():
    callback = FakeCallback(user_id=23, data="engineer_task_edit_read_only")

    result = asyncio.run(
        handle_engineer_task_callback(callback, config=TelegramConfig(owner_id=23))
    )

    assert result is not None
    assert "исправленный текст задачи" in result
    assert "read-only" in result
    assert callback.answers == [{"text": result, "show_alert": False}]
    assert callback.message.documents == []


def test_engineer_task_callback_blocks_non_owner_and_unknown_action():
    non_owner = FakeCallback(user_id=24, data="engineer_task_allow_read_only")
    unknown = FakeCallback(user_id=25, data="unknown")

    non_owner_result = asyncio.run(
        handle_engineer_task_callback(non_owner, config=TelegramConfig(owner_id=999))
    )
    unknown_result = asyncio.run(
        handle_engineer_task_callback(unknown, config=TelegramConfig(owner_id=25))
    )

    assert non_owner_result is None
    assert non_owner.answers == [
        {"text": "Доступ закрыт. Это личный бот владельца.", "show_alert": True}
    ]
    assert non_owner.message.documents == []
    assert unknown_result is not None
    assert "Неизвестная кнопка" in unknown_result
    assert "Hermes/Codex не запускаются" in unknown_result
    assert unknown.answers == [{"text": unknown_result, "show_alert": True}]
    assert unknown.message.documents == []


class FakeMessage:
    def __init__(self):
        self.documents = []
        self.texts = []

    async def answer(self, text, **kwargs):
        self.texts.append(text)

    async def answer_document(self, document, *, caption=None):
        self.documents.append({"document": document, "caption": caption})


class FakeCallback:
    def __init__(self, *, user_id: int, data: str = "download_corel_excel"):
        self.from_user = SimpleNamespace(id=user_id)
        self.data = data
        self.message = FakeMessage()
        self.answers = []

    async def answer(self, text, *, show_alert=False):
        self.answers.append({"text": text, "show_alert": show_alert})


def test_disputed_order_clears_corel_excel_context(tmp_path):
    store = InMemoryModeSessionStore()
    build_live_text_response(text="/заказ", user_id=11, session_store=store)
    build_live_text_response(text="1000*400", user_id=11, session_store=store)
    build_live_text_response(
        text="1000*400\nнепонятная строка",
        user_id=11,
        session_store=store,
    )

    try:
        prepare_corel_excel_for_user(user_id=11, output_dir=tmp_path)
    except ValueError as error:
        assert "Нет подготовленного чистого заказа" in str(error)
    else:
        raise AssertionError("disputed order must not prepare Corel Excel")


def test_order_mode_disputed_text_keeps_fix_copy_button_live_path():
    store = InMemoryModeSessionStore()
    build_live_text_response(text="/заказ", user_id=1, session_store=store)

    response = build_live_text_response(
        text="1000*400\nнепонятная строка\n1000*600",
        user_id=1,
        session_store=store,
    )

    assert response.copy_keyboard is not None
    labels = [button.label for button in response.copy_keyboard.buttons]
    assert " Скопировать для Corel" not in labels
    assert "Скачать Файл Малярки" not in labels
    assert " Скопировать подтверждённые" not in labels
    assert labels[0] == " Скопировать для исправления"
    assert response.copy_keyboard.buttons[0].copy_text == (
        "1000 400 1\nнепонятная строка\n1000 600 1"
    )


def test_chat_mode_does_not_parse_order_like_text():
    store = InMemoryModeSessionStore()
    build_live_text_response(text="/чат", user_id=2, session_store=store)

    response = build_live_text_response(
        text="1000*400",
        user_id=2,
        session_store=store,
    )

    assert store.get_mode(2) == TelegramMode.CHAT
    assert "Предпросмотр заказа" not in response.text
    assert "1. 1000 400 1" not in response.text
    assert "/заказ" in response.text
    assert "Скачать Excel для Corel" not in response.text
    assert response.copy_keyboard is None


def test_neutral_mode_parses_order_like_text_without_mode_buttons():
    store = InMemoryModeSessionStore()

    response = build_live_text_response(
        text="1000*400",
        user_id=3,
        session_store=store,
    )

    assert store.get_mode(3) == TelegramMode.ORDER
    assert "Предпросмотр заказа" in response.text
    assert "1. 1000 400 1" in response.text
    assert "/заказ" not in response.text
    assert response.copy_keyboard is not None
    assert [button.label for button in response.copy_keyboard.buttons] == [
        "Скачать Excel для Corel",
        "Скачать Файл Малярки",
        " Скопировать для Corel",
    ]


def test_engineer_task_card_and_other_read_only_modes_are_safe():
    store = InMemoryModeSessionStore()

    build_live_text_response(
        text="/инженер",
        user_id=4,
        session_store=store,
        config=TelegramConfig(owner_id=4),
    )
    engineer = build_live_text_response(
        text="подготовь задачу",
        user_id=4,
        session_store=store,
        config=TelegramConfig(owner_id=4),
    )
    assert "ЗАДАЧА" in engineer.text
    assert "подготовь задачу" in engineer.text
    assert "ОБЯЗАТЕЛЬНЫЕ ПРОВЕРКИ" in engineer.text
    assert "НЕЗАВИСИМОЕ РЕВЬЮ" in engineer.text
    assert "read-only / copy-ready" in engineer.text
    assert engineer.copy_keyboard is not None
    assert [button.label for button in engineer.copy_keyboard.buttons] == [
        "Разрешаю",
        "Отмена",
        "Исправить задачу",
    ]
    assert all(button.copy_text is None for button in engineer.copy_keyboard.buttons)

    build_live_text_response(text="/идеи", user_id=4, session_store=store, config=TelegramConfig(owner_id=4))
    ideas = build_live_text_response(text="1000*400", user_id=4, session_store=store, config=TelegramConfig(owner_id=4))
    assert "Режим идей включён" in ideas.text
    assert "Предпросмотр заказа" not in ideas.text
    assert ideas.copy_keyboard is None

    build_live_text_response(text="/выжимка", user_id=4, session_store=store)
    summary = build_live_text_response(text="идея", user_id=4, session_store=store)
    assert "выжимки" in summary.text
    assert summary.copy_keyboard is None

    build_live_text_response(text="/правила", user_id=4, session_store=store)
    rules = build_live_text_response(text="размеры", user_id=4, session_store=store)
    assert "правил" in rules.text
    assert rules.copy_keyboard is None


def test_owner_config_blocks_non_owner_live_text_and_excel_callback(tmp_path):
    store = InMemoryModeSessionStore()
    config = TelegramConfig(owner_id=100)

    response = build_live_text_response(
        text="/заказ",
        user_id=200,
        session_store=store,
        config=config,
    )
    assert "Доступ закрыт" in response.text
    assert store.get_mode(200) == TelegramMode.NEUTRAL

    callback = FakeCallback(user_id=200)
    result = asyncio.run(
        handle_corel_excel_callback(callback, output_dir=tmp_path, config=config)
    )
    assert result is None
    assert callback.message.documents == []
    assert callback.answers == [
        {"text": "Доступ закрыт. Это личный бот владельца.", "show_alert": True}
    ]


def test_app_router_integration_does_not_reference_ai_or_database():
    source = Path("malyarka_telegram/app.py").read_text(encoding="utf-8")

    assert "malyarka_hermes" not in source
    assert "deepseek" not in source.lower()
    assert "DEEPSEEK_API_KEY" not in source
    assert "GEMINI_API_KEY" not in source
    assert "orders.db" not in source
    assert "sqlite" not in source.lower()
    assert "gateway" not in source.lower()


def test_app_router_source_keeps_polling_behind_explicit_cli_flag():
    source = Path("malyarka_telegram/app.py").read_text(encoding="utf-8")

    assert "--run-polling" in source
    assert "if args.run_polling:" in source
    before_polling_flag = source.split("if args.run_polling:")[0]
    assert "run_polling()" not in before_polling_flag


def test_old_handle_text_message_still_works_without_routing_arguments():
    message = handle_text_message("1000*400")

    assert "Предпросмотр заказа" in message
    assert "1. 1000 400 1" in message


def test_clarify_intent_callback_order_switches_mode():
    callback = FakeCallback(user_id=30, data="free_entry_order")

    result = asyncio.run(
        handle_clarify_intent_callback(callback, config=TelegramConfig(owner_id=30))
    )

    assert result == "mode_switched_to_order"
    assert callback.answers == [
        {"text": "Включён режим /заказ", "show_alert": False}
    ]
    store = get_runtime_session_store()
    assert store.get_mode(30) == TelegramMode.ORDER
    store.reset_mode(30)


def test_clarify_intent_callback_cancel_stays_neutral():
    callback = FakeCallback(user_id=31, data="free_entry_cancel")

    result = asyncio.run(
        handle_clarify_intent_callback(callback, config=TelegramConfig(owner_id=31))
    )

    assert result == "clarification_cancelled"
    assert callback.answers == [
        {"text": "Уточнение отменено", "show_alert": False}
    ]
    store = get_runtime_session_store()
    assert store.get_mode(31) == TelegramMode.NEUTRAL


def test_clarify_intent_callback_ideas_switches_mode():
    callback = FakeCallback(user_id=32, data="free_entry_ideas")

    result = asyncio.run(
        handle_clarify_intent_callback(callback, config=TelegramConfig(owner_id=32))
    )

    assert result == "mode_switched_to_ideas"
    assert callback.answers == [
        {"text": "Включён режим /идеи", "show_alert": False}
    ]
    store = get_runtime_session_store()
    assert store.get_mode(32) == TelegramMode.IDEAS
    store.reset_mode(32)


def test_clarify_intent_callback_engineer_switches_mode():
    callback = FakeCallback(user_id=33, data="free_entry_engineer")

    result = asyncio.run(
        handle_clarify_intent_callback(callback, config=TelegramConfig(owner_id=33))
    )

    assert result == "mode_switched_to_engineer"
    assert callback.answers == [
        {"text": "Включён режим /инженер", "show_alert": False}
    ]
    store = get_runtime_session_store()
    assert store.get_mode(33) == TelegramMode.ENGINEER
    store.reset_mode(33)


def test_clarify_intent_callback_blocks_non_owner():
    non_owner = FakeCallback(user_id=34, data="free_entry_order")

    result = asyncio.run(
        handle_clarify_intent_callback(non_owner, config=TelegramConfig(owner_id=999))
    )

    assert result is None
    assert non_owner.answers == [
        {"text": "Доступ закрыт. Это личный бот владельца.", "show_alert": True}
    ]


def test_clarify_intent_callback_unknown_action_shows_alert():
    callback = FakeCallback(user_id=35, data="free_entry_unknown")

    result = asyncio.run(
        handle_clarify_intent_callback(callback, config=TelegramConfig(owner_id=35))
    )

    assert result is None
    assert callback.answers == [
        {"text": "Неизвестная кнопка", "show_alert": True}
    ]
