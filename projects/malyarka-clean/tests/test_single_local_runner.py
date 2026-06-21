from pathlib import Path

from openpyxl import load_workbook

from run_local_order import run_local_order


def test_clean_order_saves_text_result_and_creates_excel(tmp_path):
    input_file = _write_input(tmp_path, "1000 400 2\n700 300")
    excel_file = tmp_path / "COREL_EXPORT.xlsx"

    result = run_local_order(input_file, excel_file)

    assert result["excel_created"] is True
    assert excel_file.exists()
    assert Path(result["text_result_saved"]).exists()
    assert "Статус: чистый заказ (clean)" in Path(result["text_result_saved"]).read_text(encoding="utf-8")


def test_disputed_order_saves_text_result_and_does_not_update_existing_excel(tmp_path):
    input_file = _write_input(tmp_path, "1000 400 2\nмусор")
    excel_file = tmp_path / "COREL_EXPORT.xlsx"
    excel_file.write_bytes(b"old excel placeholder")
    before = excel_file.read_bytes()

    result = run_local_order(input_file, excel_file)

    assert result["excel_created"] is False
    assert result["block_reason"] == "disputed_rows_present"
    assert excel_file.read_bytes() == before
    text = Path(result["text_result_saved"]).read_text(encoding="utf-8")
    assert "Статус: есть спорные строки (has_disputes)" in text
    assert "Причина: есть спорные строки" in text


def test_empty_or_invalid_order_saves_text_result_and_does_not_update_existing_excel(tmp_path):
    input_file = _write_input(tmp_path, "привет\nничего непонятно")
    excel_file = tmp_path / "COREL_EXPORT.xlsx"
    excel_file.write_bytes(b"old excel placeholder")
    before = excel_file.read_bytes()

    result = run_local_order(input_file, excel_file)

    assert result["excel_created"] is False
    assert result["block_reason"] == "empty_or_invalid"
    assert excel_file.read_bytes() == before
    text = Path(result["text_result_saved"]).read_text(encoding="utf-8")
    assert "Статус: пустой или непонятный заказ (empty_or_invalid)" in text
    assert "Причина: пустой или непонятный заказ" in text


def test_clean_runner_excel_structure(tmp_path):
    input_file = _write_input(tmp_path, "1000 400 2")
    excel_file = tmp_path / "COREL_EXPORT.xlsx"

    run_local_order(input_file, excel_file)

    sheet = load_workbook(excel_file).active
    assert [sheet.cell(row=1, column=column).value for column in range(1, 4)] == [None, None, None]
    assert [sheet.cell(row=2, column=column).value for column in range(1, 4)] == [1000, 400, 2]
    assert sheet.cell(row=2, column=4).value is None


def _write_input(tmp_path, text):
    input_file = tmp_path / "ORDER_INPUT.txt"
    input_file.write_text(text, encoding="utf-8")
    return input_file
