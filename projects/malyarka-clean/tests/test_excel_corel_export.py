from pathlib import Path

from openpyxl import load_workbook

from malyarka_clean_core import build_order_result, create_corel_excel


def test_clean_order_creates_corel_export_xlsx(tmp_path):
    output_path = tmp_path / "COREL_EXPORT.xlsx"
    order_result = build_order_result("1000 400 2\n700 300")

    result = create_corel_excel(order_result, output_path)

    assert result["created"] is True
    assert result["export_blocked"] is False
    assert result["reason"] == "created"
    assert output_path.exists()


def test_corel_export_has_empty_first_row_and_no_headers(tmp_path):
    output_path = tmp_path / "COREL_EXPORT.xlsx"
    order_result = build_order_result("1000 400 2")

    create_corel_excel(order_result, output_path)

    sheet = load_workbook(output_path).active
    assert [sheet.cell(row=1, column=column).value for column in range(1, 4)] == [None, None, None]
    assert [sheet.cell(row=2, column=column).value for column in range(1, 4)] == [1000, 400, 2]


def test_corel_export_contains_only_height_width_quantity(tmp_path):
    output_path = tmp_path / "COREL_EXPORT.xlsx"
    order_result = build_order_result("1000 400 2\n700 300")

    create_corel_excel(order_result, output_path)

    sheet = load_workbook(output_path).active
    assert [sheet.cell(row=2, column=column).value for column in range(1, 4)] == [1000, 400, 2]
    assert [sheet.cell(row=3, column=column).value for column in range(1, 4)] == [700, 300, 1]
    assert sheet.cell(row=2, column=4).value is None
    assert sheet.cell(row=3, column=4).value is None


def test_disputed_order_blocks_export_and_does_not_create_file(tmp_path):
    output_path = tmp_path / "COREL_EXPORT.xlsx"
    order_result = build_order_result("1000 400 2\nмусор")

    result = create_corel_excel(order_result, output_path)

    assert result["created"] is False
    assert result["export_blocked"] is True
    assert result["reason"] == "disputed_rows_present"
    assert result["source_status"] == "has_disputes"
    assert not output_path.exists()


def test_empty_or_invalid_order_blocks_export_and_does_not_create_file(tmp_path):
    output_path = tmp_path / "COREL_EXPORT.xlsx"
    order_result = build_order_result("привет\nничего непонятно")

    result = create_corel_excel(order_result, output_path)

    assert result["created"] is False
    assert result["export_blocked"] is True
    assert result["reason"] == "empty_or_invalid"
    assert result["source_status"] == "empty_or_invalid"
    assert not output_path.exists()


def test_disputed_rows_are_not_written_to_excel(tmp_path):
    output_path = tmp_path / "COREL_EXPORT.xlsx"
    clean_result = build_order_result("1000 400 2")
    disputed_result = build_order_result("1000 400 2\nмусор")

    create_corel_excel(clean_result, output_path)
    before_size = Path(output_path).stat().st_size
    blocked = create_corel_excel(disputed_result, output_path)

    assert blocked["created"] is False
    assert Path(output_path).stat().st_size == before_size
    sheet = load_workbook(output_path).active
    assert sheet.cell(row=2, column=1).value == 1000
    assert sheet.cell(row=3, column=1).value is None
