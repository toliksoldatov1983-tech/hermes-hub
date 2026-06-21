"""Safe .xlsx export for Corel from already prepared Corel rows."""

from pathlib import Path

from openpyxl import Workbook

from .corel_export_model import build_corel_export_model


def create_corel_excel(order_result, output_path):
    """Create a Corel .xlsx file only when the order result is clean."""
    corel_model = build_corel_export_model(order_result)

    if corel_model["export_blocked"]:
        return {
            "created": False,
            "export_blocked": True,
            "reason": corel_model["reason"],
            "source_status": corel_model["source_status"],
            "output_path": str(output_path),
            "corel_rows": [],
        }

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    _write_corel_workbook(corel_model["corel_rows"], output)

    return {
        "created": True,
        "export_blocked": False,
        "reason": "created",
        "source_status": corel_model["source_status"],
        "output_path": str(output),
        "corel_rows": corel_model["corel_rows"],
    }


def _write_corel_workbook(corel_rows, output_path):
    workbook = Workbook()
    sheet = workbook.active

    for index, row in enumerate(corel_rows, start=2):
        sheet.cell(row=index, column=1, value=row["height_mm"])
        sheet.cell(row=index, column=2, value=row["width_mm"])
        sheet.cell(row=index, column=3, value=row["quantity"])

    workbook.save(output_path)
